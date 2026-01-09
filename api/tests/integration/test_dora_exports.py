from __future__ import annotations

import io
import json
import os
from datetime import timedelta

import django
from django.contrib.auth import get_user_model
from django.test import Client, TestCase
from django.utils import timezone
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "callcentersite.settings.testing")
django.setup()

from dora_metrics.models import DORAMetric
from dora_metrics.views import calculate_dora_classification


class DORAExportAPITest(TestCase):
    """Integration tests for DORA export endpoints and classification thresholds."""

    def setUp(self) -> None:
        self.client = Client()
        User = get_user_model()
        self.user = User.objects.create_superuser(
            username="export_admin",
            email="export@test.com",
            password="testpass123",
        )
        self.client.login(username="export_admin", password="testpass123")
        self.now = timezone.now()
        self._create_metrics()

    def _create_metrics(self) -> None:
        """Create baseline DORA metrics across two days."""
        day_1 = self.now - timedelta(days=2)
        day_2 = self.now - timedelta(days=1)

        DORAMetric.objects.create(
            cycle_id="cycle-success",
            feature_id="FEAT-OK",
            phase_name="deployment",
            decision="go",
            duration_seconds=1200,
            created_at=day_1,
        )
        DORAMetric.objects.create(
            cycle_id="cycle-failed",
            feature_id="FEAT-NO",
            phase_name="deployment",
            decision="no-go",
            duration_seconds=1800,
            created_at=day_2,
        )
        DORAMetric.objects.create(
            cycle_id="cycle-recovery",
            feature_id="FEAT-NO",
            phase_name="recovery",
            decision="resolved",
            duration_seconds=3600,
            created_at=day_2,
        )

    def test_calculate_dora_classification_thresholds(self) -> None:
        """CFR thresholds follow documented DORA ranges."""
        classification_high = calculate_dora_classification(
            deployment_count=5,
            days=30,
            lead_time_hours=30,
            cfr=30,
            mttr_hours=200,
        )
        classification_medium = calculate_dora_classification(
            deployment_count=5,
            days=30,
            lead_time_hours=30,
            cfr=31,
            mttr_hours=200,
        )

        self.assertEqual(classification_high, "High")
        self.assertEqual(classification_medium, "Medium")

    def test_export_csv_includes_metrics(self) -> None:
        """CSV export returns data rows and correct headers."""
        start = (self.now - timedelta(days=3)).date().isoformat()
        end = self.now.date().isoformat()

        response = self.client.get(
            f"/api/v1/dora/export/csv/?start_date={start}&end_date={end}"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        self.assertIn("dora_metrics", response["Content-Disposition"])

        content = response.content.decode("utf-8")
        self.assertIn("Deployment Frequency", content)
        self.assertIn("Change Failure Rate (%)", content)

    def test_export_excel_generates_workbook(self) -> None:
        """Excel export returns a valid workbook with daily metrics sheet."""
        start = (self.now - timedelta(days=3)).date().isoformat()
        end = self.now.date().isoformat()

        response = self.client.get(
            f"/api/v1/dora/export/excel/?start_date={start}&end_date={end}"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )

        workbook = load_workbook(filename=io.BytesIO(response.content))
        self.assertIn("Daily Metrics", workbook.sheetnames)
        sheet = workbook["Daily Metrics"]
        self.assertGreaterEqual(sheet.max_row, 2)

    def test_export_pdf_contains_summary(self) -> None:
        """PDF export includes a minimal PDF payload."""
        start = (self.now - timedelta(days=3)).date().isoformat()
        end = self.now.date().isoformat()

        response = self.client.get(
            f"/api/v1/dora/export/pdf/?start_date={start}&end_date={end}"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_data_catalog_endpoint_resolves(self) -> None:
        """Ensure data catalog endpoint stays versioned and functional."""
        response = self.client.get("/api/v1/dora/data-catalog/")
        self.assertEqual(response.status_code, 200)
        catalog = json.loads(response.content)
        self.assertIn("datasets", catalog)
